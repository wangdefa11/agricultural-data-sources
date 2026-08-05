#!/usr/bin/env python3
"""把 Gildata(聚源/汇易网/商务部/博亚和讯/IMF/外汇交易中心) 数据按月归集后回填采集模板。
输入: data/raw/gildata/*.csv + data/raw/wind/wind_us_soybean_planted_area.csv
输出: outputs/豆粕研究数据采集模板_月度年度_已填充2026-08-01_v2.xlsx
"""
import csv, re, glob, os
from collections import defaultdict
from datetime import datetime
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, "outputs", "豆粕研究数据采集模板_月度年度_已填充2026-08-01.xlsx")
DST = os.path.join(BASE, "outputs", "豆粕研究数据采集模板_月度年度_已填充2026-08-01_v2.xlsx")

# ---------- 1. 解析所有 gildata csv -> series[name] = {date: value}（按日期去重） ----------
def load_series(path):
    out = defaultdict(dict)
    with open(path, encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            for line in r.get("table_markdown", "").split("\n"):
                m = re.match(r"\| (\d+) \| ([^|]+) \| ([^|]+) \| ([^|]+) \| ([^|]+) \| ([^|]+) \| ([^|]+) \|", line)
                if m:
                    try:
                        v = float(m.group(5).strip())
                    except ValueError:
                        continue
                    out[m.group(2).strip()][m.group(6).strip()] = v
    return out

S = {}
for f in glob.glob(os.path.join(BASE, "data/raw/gildata/gildata_*.csv")):
    for name, d in load_series(f).items():
        S.setdefault(name, {}).update(d)

def month_of(d):  # '2026-07-31' -> '2026-07'
    return d[:7]

def mavg(name, drop_zero=True):
    """月均值"""
    acc = defaultdict(list)
    for d, v in S.get(name, {}).items():
        if drop_zero and v == 0:
            continue
        acc[month_of(d)].append(v)
    return {m: sum(vs) / len(vs) for m, vs in acc.items()}

def mlast(name, drop_zero=True):
    """月末值（当月最后一个观测）"""
    obs = sorted((d, v) for d, v in S.get(name, {}).items() if not (drop_zero and v == 0))
    out = {}
    for d, v in obs:
        out[month_of(d)] = v  # 有序遍历，最后覆盖即月末
    return out

def msum(name, min_parts=2):
    """半月->月度合计，返回 (值, 份数)"""
    acc = defaultdict(list)
    for d, v in S.get(name, {}).items():
        acc[month_of(d)].append(v)
    return {m: (sum(vs), len(vs)) for m, vs in acc.items()}

# 月度归集
spot_meal   = mavg("中国:现货价:豆粕:当期值:日")                 # 元/吨 2026-05~07
soyoil      = mavg("中国:平均价:豆油:当期值:日")                 # 元/吨
corn        = mavg("中国:现货平均价:玉米:当期值:日")             # 元/吨
fx_m        = {m[:7]: v for m, v in S.get("平均汇率:美元兑人民币:当期值:月", {}).items()}
cbot_m      = {m[:7]: v for m, v in S.get("环球:期货价格:大豆:芝加哥大豆:当期值:月", {}).items()}
port_stocks = {m: v / 10000 for m, v in mlast("中国:商业库存量:大豆:期末值:日").items()}   # 吨->万吨
meal_stocks = mlast("中国:商业库存量:豆粕:期末值:周")            # 万吨
arrivals    = msum("中国:抵港数量:大豆:当期值:半月")             # 吨, 半月
hog_profit  = mavg("中国:养殖利润:自繁生猪:当期值:周")           # 元/头
broiler_p   = mavg("中国:屠宰利润:白羽肉鸡:当期值:周")           # 元/羽
layer_p     = mavg("中国:养殖利润:蛋鸡:当期值:周")               # 元/羽
landed = {src: mavg(f"{src}:到岸完税价格:大豆:当期值:日")
          for src in ["巴西", "美国美湾", "美国西部地区", "阿根廷"]}
margin_tj = mavg("天津:压榨利润:大豆:进口:当期值:日")
margin_rz = mavg("山东:日照:压榨利润:大豆:进口:当期值:日")
import_cny_wy = {m[:7]: v for m, v in S.get("中国:进口金额:以人民币计:大豆:当期值:月", {}).items()}  # 万元

# Wind NASS 播种面积（千英亩）-> 千公顷
nass = {}
with open(os.path.join(BASE, "data/raw/wind/wind_us_soybean_planted_area.csv"), encoding="utf-8") as fh:
    for r in csv.DictReader(fh):
        nass[int(r["日期"][:4])] = float(r["美国:播种面积:大豆"]) * 0.404686

def ym(dt): return f"{dt.year:04d}-{dt.month:02d}"
def dt_of(m): return datetime(int(m[:4]), int(m[5:7]), 1)

# ---------- 2. 打开工作簿 ----------
wb = openpyxl.load_workbook(SRC)
stats = defaultdict(int)

# ---- 02 大豆主产国年度：美国播种面积（千公顷，NASS 换算）----
ws = wb["02大豆主产国年度"]
for r in range(5, ws.max_row + 1):
    my, country = ws.cell(r, 1).value, ws.cell(r, 3).value
    if country == "美国" and my:
        yr = int(str(my)[:4])
        if yr in nass and ws.cell(r, 4).value in (None, ""):
            ws.cell(r, 4, round(nass[yr], 0))
            note = ws.cell(r, 13).value or ""
            ws.cell(r, 13, note + "；播种面积=NASS 千英亩×0.4047 换算（Wind/NASS）")
            stats["02播种面积"] += 1

# ---- 03 中国进口月度：追加“全部来源国(合计)”行（数量=分国加总，金额=海关人民币亿元）----
ws = wb["03中国进口月度"]
qty_by_month = defaultdict(float)
for r in range(5, ws.max_row + 1):
    a, c = ws.cell(r, 1).value, ws.cell(r, 3).value
    if a is not None and c is not None:
        qty_by_month[ym(a)] += float(c)
customs_url = next(ws.cell(r, 6).value for r in range(5, 10) if ws.cell(r, 6).value)
r = ws.max_row + 1
for m in sorted(import_cny_wy):
    if m not in qty_by_month:
        continue
    ws.cell(r, 1, dt_of(m))
    ws.cell(r, 2, "全部来源国(合计)")
    ws.cell(r, 3, round(qty_by_month[m], 2))
    ws.cell(r, 4, round(import_cny_wy[m] / 10000, 2))          # 万元->亿元
    ws.cell(r, 5, f'=IF(OR(C{r}="",D{r}=""),"",D{r}*10000/C{r})')
    ws.cell(r, 6, customs_url)
    ws.cell(r, 7, "金额=海关总署人民币口径（Gildata 采集，万元→亿元）；数量=表内分来源国加总")
    r += 1
    stats["03合计行"] += 1

# ---- 04 压榨库存月度：到港量 / 港口大豆库存 / 油厂豆粕库存 ----
ws = wb["04压榨库存月度"]
for r in range(5, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a is None:
        continue
    m = ym(a)
    got = []
    if m in arrivals and arrivals[m][1] >= 2 and ws.cell(r, 3).value in (None, ""):
        ws.cell(r, 3, round(arrivals[m][0] / 10000, 2))        # 吨->万吨
        got.append("到港=商务部半月合计")
    if m in port_stocks and ws.cell(r, 6).value in (None, ""):
        ws.cell(r, 6, round(port_stocks[m], 2))
        got.append("港口大豆库存=汇易网月末值")
    if m in meal_stocks and ws.cell(r, 8).value in (None, ""):
        ws.cell(r, 8, round(meal_stocks[m], 2))
        got.append("油厂豆粕库存=汇易网周度月末值")
    if got:
        note = ws.cell(r, 13).value or ""
        ws.cell(r, 13, note + "；" + "、".join(got) + "（Gildata 2026-08-01 采集）")
        stats["04行"] += 1

# ---- 07 价格基差月度：补 CBOT/豆油/玉米/汇率，并向历史延展 ----
ws = wb["07价格基差月度"]
existing = {}
for r in range(5, ws.max_row + 1):
    if ws.cell(r, 1).value is not None:
        existing[ym(ws.cell(r, 1).value)] = r
for m, r in existing.items():
    got = []
    if m in cbot_m and ws.cell(r, 7).value in (None, ""):
        ws.cell(r, 7, round(cbot_m[m], 2)); got.append("CBOT=IMF月均(美元/吨)")
    if m in soyoil and ws.cell(r, 8).value in (None, ""):
        ws.cell(r, 8, round(soyoil[m], 2)); got.append("豆油=汇易网月均")
    if m in corn and ws.cell(r, 9).value in (None, ""):
        ws.cell(r, 9, round(corn[m], 2)); got.append("玉米=汇易网月均")
    if m in fx_m and ws.cell(r, 10).value in (None, ""):
        ws.cell(r, 10, round(fx_m[m], 4)); got.append("汇率=外汇交易中心月均")
    if got:
        note = ws.cell(r, 13).value or ""
        ws.cell(r, 13, note + "；" + "、".join(got) + "（Gildata）；DCE期货价/海运费暂无公开月度源")
        stats["07补列"] += 1
# 历史延展行：2023-08 ~ 2026-03（CBOT+汇率均有）
months_hist = []
for y in range(2023, 2027):
    for mm in range(1, 13):
        m = f"{y:04d}-{mm:02d}"
        if "2023-08" <= m <= "2026-03":
            months_hist.append(m)
r = ws.max_row + 1
# 先用模板内剩余空行（带公式），不够再追加
free_rows = [rr for rr in range(5, ws.max_row + 1) if ws.cell(rr, 1).value is None]
idx = 0
for m in months_hist:
    if idx < len(free_rows):
        rr = free_rows[idx]; idx += 1
    else:
        rr = ws.max_row + 1
        ws.cell(rr, 6, f'=IF(OR(C{rr}="",E{rr}=""),"",E{rr}-C{rr})')
    ws.cell(rr, 1, dt_of(m))
    ws.cell(rr, 2, "全国均价")
    if m in cbot_m: ws.cell(rr, 7, round(cbot_m[m], 2))
    if m in fx_m:   ws.cell(rr, 10, round(fx_m[m], 4))
    ws.cell(rr, 12, "Gildata（IMF/中国外汇交易中心，2026-08-01 采集）")
    ws.cell(rr, 13, "仅 CBOT 大豆月均(美元/吨,IMF)与人民币月均汇率；现货/期货/豆油/玉米/海运暂无该月公开数据")
    stats["07历史行"] += 1

# ---- 08 成本利润月度：2026-05~07 × 4 来源 + 2025-08~2026-04 仅养殖利润 ----
ws = wb["08成本利润月度"]
def fill08(rr, m, src=None):
    if src is not None:
        ws.cell(rr, 1, dt_of(m)); ws.cell(rr, 2, src)
        ws.cell(rr, 3, round(landed[src][m], 2))
        ws.cell(rr, 4, round(spot_meal[m], 2))
        ws.cell(rr, 5, round(soyoil[m], 2))
        ws.cell(rr, 6, 0.79); ws.cell(rr, 7, 0.185); ws.cell(rr, 8, 130)
        ws.cell(rr, 14, "Gildata/汇易网（2026-08-01 采集）")
        ws.cell(rr, 15, "得率(粕79%/油18.5%)与加工费为行业假设(加工费=Wind国产大豆压榨成本130元/吨代理)；"
                        f"交叉验证:汇易网压榨利润月均 天津{margin_tj.get(m, float('nan')):.0f}/日照{margin_rz.get(m, float('nan')):.0f} 元/吨")
    else:
        ws.cell(rr, 1, dt_of(m)); ws.cell(rr, 2, "—（仅养殖利润）")
        ws.cell(rr, 14, "Gildata/博亚和讯、涌益（2026-08-01 采集）")
        ws.cell(rr, 15, "仅养殖利润（周度→月均）；到岸成本/销售价暂无该月公开数据")
    if m in hog_profit: ws.cell(rr, 11, round(hog_profit[m], 1))
    if m in broiler_p:  ws.cell(rr, 12, round(broiler_p[m], 2))
    if m in layer_p:    ws.cell(rr, 13, round(layer_p[m], 2))
    if rr > 24:  # 超出模板预置行，补公式
        ws.cell(rr, 9,  f'=IF(OR(D{rr}="",E{rr}="",F{rr}="",G{rr}=""),"",D{rr}*F{rr}+E{rr}*G{rr})')
        ws.cell(rr, 10, f'=IF(OR(I{rr}="",C{rr}="",H{rr}=""),"",I{rr}-C{rr}-H{rr})')

free_rows = [rr for rr in range(5, ws.max_row + 1) if ws.cell(rr, 1).value is None]
fi = 0
def next_row():
    global fi
    if fi < len(free_rows):
        rr = free_rows[fi]; fi += 1; return rr
    return ws.max_row + 1

for m in ["2026-05", "2026-06", "2026-07"]:
    for src in ["巴西", "美国美湾", "美国西部地区", "阿根廷"]:
        if m in landed[src] and m in spot_meal and m in soyoil:
            fill08(next_row(), m, src)
            stats["08成本行"] += 1
for y in range(2025, 2027):
    for mm in range(1, 13):
        m = f"{y:04d}-{mm:02d}"
        if "2025-08" <= m <= "2026-04":
            fill08(next_row(), m)
            stats["08利润行"] += 1

# ---- 数据源与口径：补充 Gildata 渠道说明 ----
ws = wb["数据源与口径"]
r = ws.max_row + 1
rows = [
    ["聚源 Gildata（汇易网）", "豆粕/豆油/玉米现货价、港口与油厂库存、分港压榨利润、分来源到岸完税价", "日/周→月度", "元/吨、万吨", "月均/月末", "Gildata 插件采集 2026-08-01"],
    ["聚源 Gildata（商务部/博亚和讯/涌益/IMF/外汇交易中心）", "大豆抵港半月、养殖利润、CBOT 大豆月均、人民币月均汇率", "半月/周/月", "吨、元/头、美元/吨", "月合计/月均", "Gildata 插件采集 2026-08-01"],
]
for row in rows:
    for c, v in enumerate(row, 1):
        ws.cell(r, c, v)
    r += 1
    stats["口径说明"] += 1

wb.save(DST)
print("saved:", DST)
for k, v in stats.items():
    print(f"  {k}: {v}")
