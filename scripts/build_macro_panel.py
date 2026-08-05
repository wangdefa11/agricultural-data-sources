"""宏观展示面板：把 wind_observations 统一到月度/年度粒度。

产出：
- SQLite 表 macro_monthly(month, indicator, value, note)
- Excel 工作簿 outputs/农产品宏观面板_月度.xlsx（月度面板 + 年度面板 + 数据说明）

粒度规则：日度 -> 月均值；月度 -> 原值；年度 -> 年度面板。
"""
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "data" / "soybean.sqlite3"
OUT = ROOT / "outputs" / "农产品宏观面板_月度.xlsx"

MONTHLY_INDICATORS = {
    "中国:现货价:豆粕": ("豆粕现货价月均", "元/吨", "日度现货价按月平均"),
    "中国:压榨量:大豆#2": ("大豆压榨量(实际月度)", "万吨/月", "Wind 实际月度列"),
    "中国:压榨量:大豆": ("大豆压榨量(年化均摊)", "万吨/月", "年化值/12，仅作参照"),
}
ANNUAL_INDICATORS = [
    "中国:库存量:豆粕",
    "中国:期初库存量:豆粕",
    "中国:期末库存量:豆粕",
    "中国:年末存栏数量:父母代肉鸡场",
    "中国:年末存栏数量:祖代及以上肉鸡场",
    "北京:畜禽存栏量:家禽:产蛋鸡",
    "天津:年末存栏量:家禽:产蛋鸡",
    "山东:济宁:存栏量:活家禽:活鸡:蛋鸡",
    "辽宁:大连:期末存栏数:家禽:蛋鸡",
    "辽宁:沈阳:畜牧业:年末存栏数:家禽:蛋鸡",
]


def main() -> None:
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS macro_monthly")
    cur.execute(
        "CREATE TABLE macro_monthly(month TEXT, indicator TEXT, value REAL, note TEXT, "
        "PRIMARY KEY(month, indicator))"
    )

    for ind, (name, unit, note) in MONTHLY_INDICATORS.items():
        rows = cur.execute(
            "SELECT substr(date,1,7) ym, AVG(value) FROM wind_observations "
            "WHERE indicator=? GROUP BY ym", (ind,),
        ).fetchall()
        for ym, v in rows:
            cur.execute(
                "INSERT OR REPLACE INTO macro_monthly VALUES(?,?,?,?)",
                (ym, f"{name}({unit})", round(v, 2), note),
            )

    # 年度面板直接从 wind_observations 取（不建库表，只进 Excel）
    annual: dict[str, dict[str, float]] = {}
    for ind in ANNUAL_INDICATORS:
        for date, v in cur.execute(
            "SELECT date, value FROM wind_observations WHERE indicator=?", (ind,)
        ):
            year = date[:4]
            annual.setdefault(year, {})[ind] = v

    # CBOT 快照说明
    cbot = cur.execute(
        "SELECT contract, close, unit FROM wind_cbot_snapshot WHERE close>0 ORDER BY contract"
    ).fetchall()
    con.commit()

    months = [r[0] for r in cur.execute(
        "SELECT DISTINCT month FROM macro_monthly ORDER BY month")]
    mind = [r[0] for r in cur.execute(
        "SELECT DISTINCT indicator FROM macro_monthly ORDER BY indicator")]
    panel = {(m, i): v for m, i, v in cur.execute(
        "SELECT month, indicator, value FROM macro_monthly")}

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "月度面板"
    ws.append(["月份"] + mind)
    for c in ws[1]:
        c.font = bold
    for m in months:
        ws.append([m] + [panel.get((m, i)) for i in mind])

    ws2 = wb.create_sheet("年度面板")
    years = sorted(annual, reverse=True)
    ws2.append(["年份"] + ANNUAL_INDICATORS)
    for c in ws2[1]:
        c.font = bold
    for y in years:
        ws2.append([y] + [annual[y].get(i) for i in ANNUAL_INDICATORS])

    ws3 = wb.create_sheet("数据说明")
    notes = [
        ["生成方式", "scripts/build_macro_panel.py，源表 wind_observations（scripts/load_wind_data.py 入库）"],
        ["粒度", "宏观展示用：日度已按月平均，月度原值保留，年度单独成表"],
        ["豆粕现货价", "Wind 日度现货(元/吨)按月平均；日度原始覆盖 2026-04 起"],
        ["大豆压榨量", "实际月度=Wind 月度列；年化均摊列×12 与 USDA 市场年度值一致"],
        ["CBOT 大豆", "Wind 仅提供最新合约快照(2026-07-31)，月度历史暂不可得"],
        ["CBOT 快照", "; ".join(f"{c}={p}{u}" for c, p, u in cbot[:6])],
        ["蛋鸡存栏", "Wind 仅覆盖北京/天津/济宁/大连/沈阳，非全国完整面板"],
        ["口径提醒", "月度压榨(日历年)与 USDA 市场年度值存在口径差，勿无说明拼接"],
    ]
    for row in notes:
        ws3.append(row)
    for c in ws3[1]:
        pass
    ws3["A1"].font = bold
    ws3["B1"].font = bold

    for sheet in (ws, ws2, ws3):
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 24
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"macro_monthly {cur.execute('SELECT COUNT(*) FROM macro_monthly').fetchone()[0]} 行")
    print(f"Excel: {OUT}")
    con.close()


if __name__ == "__main__":
    main()
