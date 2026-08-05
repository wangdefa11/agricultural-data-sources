from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "outputs"
OUTPUT_PATH = OUTPUT_DIR / "豆粕研究数据采集模板.xlsx"
MAX_ROWS = 24

RED = "D71900"
PALE_RED = "FCE4DD"
INPUT_FILL = "FFF2CC"
FORMULA_FILL = "E2F0D9"
DARK = "222222"
BLUE = "0070C0"
WHITE = "FFFFFF"
GRID = "D9D9D9"
FONT_NAME = "Microsoft YaHei"


SCHEMAS = [
    {
        "name": "01全球油籽年度",
        "title": "全球油籽供需｜年度/营销年度",
        "frequency": "年度",
        "source": "USDA PSD",
        "source_url": "https://apps.fas.usda.gov/PSDOnlineDataServices/swagger/ui/index",
        "fields": [
            ("marketing_year", "营销年度", "文本", "YYYY/YY，例如2025/26", "必填", "原始"),
            ("version_date", "数据版本日期", "日期", "yyyy-mm-dd", "必填", "原始"),
            ("commodity", "油籽品种", "文本", "大豆/菜籽/葵花籽/棉籽/花生/椰肉干/棕榈仁", "必填", "原始"),
            ("country_region", "国家或地区", "文本", "World或国家名", "必填", "原始"),
            ("production_kt", "产量", "数值", "千吨", "必填", "原始"),
            ("exports_kt", "出口量", "数值", "千吨", "必填", "原始"),
            ("imports_kt", "进口量", "数值", "千吨", "必填", "原始"),
            ("crush_kt", "压榨量", "数值", "千吨", "必填", "原始"),
            ("domestic_use_kt", "国内消费量", "数值", "千吨", "选填", "原始"),
            ("ending_stocks_kt", "期末库存", "数值", "千吨", "选填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL", "必填", "原始"),
            ("note", "备注", "文本", "口径变化、预测/估计说明", "选填", "原始"),
        ],
    },
    {
        "name": "02大豆主产国年度",
        "title": "大豆主产国生产与平衡｜年度/营销年度",
        "frequency": "年度",
        "source": "USDA PSD、USDA NASS、CONAB、阿根廷农业部",
        "source_url": "https://apps.fas.usda.gov/PSDOnline/app/index.html",
        "fields": [
            ("marketing_year", "营销年度", "文本", "YYYY/YY", "必填", "原始"),
            ("version_date", "数据版本日期", "日期", "yyyy-mm-dd", "必填", "原始"),
            ("country", "国家", "文本", "美国/巴西/阿根廷/中国/其他", "必填", "原始"),
            ("planted_area_kha", "播种面积", "数值", "千公顷", "必填", "原始"),
            ("harvested_area_kha", "收获面积", "数值", "千公顷", "选填", "原始"),
            ("yield_t_ha", "单产", "数值", "吨/公顷", "必填", "原始"),
            ("production_kt", "产量", "数值", "千吨", "必填", "原始"),
            ("exports_kt", "出口量", "数值", "千吨", "选填", "原始"),
            ("imports_kt", "进口量", "数值", "千吨", "选填", "原始"),
            ("crush_kt", "压榨量", "数值", "千吨", "选填", "原始"),
            ("ending_stocks_kt", "期末库存", "数值", "千吨", "选填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL", "必填", "原始"),
            ("note", "备注", "文本", "机构口径及修订说明", "选填", "原始"),
        ],
    },
    {
        "name": "03中国进口月度",
        "title": "中国大豆进口｜月度",
        "frequency": "月度",
        "source": "海关总署",
        "source_url": "https://english.customs.gov.cn/Statistics/Statistics",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("origin_country", "来源国", "文本", "巴西/美国/阿根廷/其他/合计", "必填", "原始"),
            ("import_qty_10kt", "进口数量", "数值", "万吨", "必填", "原始"),
            ("import_value_100m_cny", "进口金额", "数值", "亿元人民币", "选填", "原始"),
            ("avg_import_price_cny_t", "进口均价", "数值", "元/吨；金额×10000÷数量", "自动", "公式"),
            ("source_url", "来源网址", "文本", "完整URL", "必填", "原始"),
            ("note", "备注", "文本", "当月值/累计值必须注明", "选填", "原始"),
        ],
        "formulas": {5: lambda row: f'=IF(OR(C{row}="",D{row}=""),"",D{row}*10000/C{row})'},
    },
    {
        "name": "04压榨库存月度",
        "title": "中国压榨与库存｜月度",
        "frequency": "月度",
        "source": "行业资讯、粮油信息中心、港口及油厂样本",
        "source_url": "",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("region", "区域", "文本", "全国/华东/华南/华北/山东/东北等", "必填", "原始"),
            ("soy_arrivals_10kt", "大豆到港量", "数值", "万吨；月度合计", "必填", "原始"),
            ("crush_volume_10kt", "大豆压榨量", "数值", "万吨；月度合计", "必填", "原始"),
            ("operating_rate_pct", "平均开机率", "百分比", "0—100%", "选填", "原始"),
            ("port_soy_stocks_10kt", "港口大豆库存", "数值", "万吨；月末值", "选填", "原始"),
            ("mill_soy_stocks_10kt", "油厂大豆库存", "数值", "万吨；月末值", "必填", "原始"),
            ("soymeal_stocks_10kt", "油厂豆粕库存", "数值", "万吨；月末值", "必填", "原始"),
            ("soymeal_output_10kt", "豆粕产量", "数值", "万吨；月度合计", "选填", "原始"),
            ("unexecuted_contracts_10kt", "豆粕未执行合同", "数值", "万吨；月末值", "选填", "原始"),
            ("source_name", "来源机构", "文本", "供应商或样本名称", "必填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL；终端数据可写终端名", "选填", "原始"),
            ("note", "备注", "文本", "样本范围及口径变化", "选填", "原始"),
        ],
    },
    {
        "name": "05饲料月度",
        "title": "饲料产量与结构｜月度",
        "frequency": "月度",
        "source": "中国饲料工业协会",
        "source_url": "https://www.chinafeed.org.cn/",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("feed_category", "饲料品种", "文本", "猪料/蛋禽料/肉禽料/水产料/反刍料/宠物料/其他/合计", "必填", "原始"),
            ("feed_output_10kt", "饲料产量", "数值", "万吨；月度合计", "必填", "原始"),
            ("yoy_pct", "同比增速", "百分比", "0.0%", "选填", "原始"),
            ("soymeal_inclusion_pct", "豆粕添加率假设", "百分比", "0.0%；若无可靠数据可暂空", "选填", "假设"),
            ("estimated_soymeal_demand_10kt", "估算豆粕需求", "数值", "万吨；饲料产量×添加率", "自动", "公式"),
            ("source_url", "来源网址", "文本", "完整URL", "必填", "原始"),
            ("note", "备注", "文本", "样本范围及口径变化", "选填", "原始"),
        ],
        "formulas": {6: lambda row: f'=IF(OR(C{row}="",E{row}=""),"",C{row}*E{row})'},
    },
    {
        "name": "06养殖月度",
        "title": "养殖存栏与供给指标｜月度",
        "frequency": "月度（季度数据统一落在季度末月份）",
        "source": "国家统计局、农业农村部、行业资料",
        "source_url": "https://www.moa.gov.cn/ztzl/szcpxx/jdsj/",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("hog_stock_10k_head", "生猪存栏", "数值", "万头；月末或季度末", "必填", "原始"),
            ("breeding_sow_10k_head", "能繁母猪存栏", "数值", "万头；月末或季度末", "必填", "原始"),
            ("hog_slaughter_10k_head", "生猪出栏/屠宰量", "数值", "万头；月度合计", "选填", "原始"),
            ("broiler_stock_10k_birds", "肉鸡存栏或补栏", "数值", "万羽；注明口径", "选填", "原始"),
            ("layer_stock_10k_birds", "蛋鸡存栏或补栏", "数值", "万羽；注明口径", "选填", "原始"),
            ("aquaculture_index", "水产养殖景气或投苗指数", "数值", "指数；没有可不采", "选填", "原始"),
            ("source_name", "来源机构", "文本", "国家统计局/农业农村部/行业机构", "必填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL", "必填", "原始"),
            ("note", "备注", "文本", "月度推算值与季度官方值需区分", "选填", "原始"),
        ],
    },
    {
        "name": "07价格基差月度",
        "title": "价格、基差与相关市场｜月度",
        "frequency": "月度",
        "source": "DCE、CME、现货报价、外汇及运费数据",
        "source_url": "https://www.dce.com.cn/",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("region", "现货区域", "文本", "全国/华东/华南/山东等", "必填", "原始"),
            ("dce_soymeal_avg_cny_t", "豆粕期货月均价", "数值", "元/吨；明确主力或指定合约", "必填", "原始"),
            ("dce_soymeal_month_end_cny_t", "豆粕期货月末价", "数值", "元/吨", "必填", "原始"),
            ("soymeal_spot_avg_cny_t", "豆粕现货月均价", "数值", "元/吨", "必填", "原始"),
            ("basis_avg_cny_t", "月均基差", "数值", "现货月均价-期货月均价", "自动", "公式"),
            ("cbot_soy_avg_uscent_bu", "CBOT大豆月均价", "数值", "美分/蒲式耳", "必填", "原始"),
            ("soybean_oil_avg_cny_t", "豆油月均价", "数值", "元/吨", "选填", "原始"),
            ("corn_avg_cny_t", "玉米月均价", "数值", "元/吨", "选填", "原始"),
            ("usd_cny_avg", "美元兑人民币月均汇率", "数值", "人民币/美元", "必填", "原始"),
            ("ocean_freight_usd_t", "海运费月均值", "数值", "美元/吨", "选填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL或行情终端名称", "必填", "原始"),
            ("note", "备注", "文本", "合约切换规则必须注明", "选填", "原始"),
        ],
        "formulas": {6: lambda row: f'=IF(OR(C{row}="",E{row}=""),"",E{row}-C{row})'},
    },
    {
        "name": "08成本利润月度",
        "title": "进口成本、压榨利润与养殖利润｜月度",
        "frequency": "月度",
        "source": "由进口、价格与行业数据计算",
        "source_url": "",
        "fields": [
            ("month", "月份", "日期", "统一填每月1日，格式yyyy-mm", "必填", "原始"),
            ("origin", "大豆来源", "文本", "美湾/巴西/阿根廷/国内", "必填", "原始"),
            ("soy_landed_cost_cny_t", "大豆到岸完税成本", "数值", "元/吨", "必填", "原始"),
            ("soymeal_price_cny_t", "豆粕销售价格", "数值", "元/吨；月均", "必填", "原始"),
            ("soyoil_price_cny_t", "豆油销售价格", "数值", "元/吨；月均", "必填", "原始"),
            ("meal_yield_pct", "豆粕得率", "百分比", "例如79.0%", "必填", "假设"),
            ("oil_yield_pct", "豆油得率", "百分比", "例如18.5%", "必填", "假设"),
            ("other_cost_cny_t", "加工及其他成本", "数值", "元/吨大豆", "必填", "假设"),
            ("crush_revenue_cny_t", "压榨收入", "数值", "豆粕价×粕得率+豆油价×油得率", "自动", "公式"),
            ("crush_margin_cny_t", "压榨利润", "数值", "压榨收入-大豆成本-其他成本", "自动", "公式"),
            ("hog_profit_cny_head", "生猪养殖利润", "数值", "元/头；月均", "选填", "原始"),
            ("broiler_profit_cny_bird", "肉鸡养殖利润", "数值", "元/羽；月均", "选填", "原始"),
            ("layer_profit_cny_bird", "蛋鸡养殖利润", "数值", "元/羽；月均", "选填", "原始"),
            ("source_url", "来源网址", "文本", "完整URL或终端名称", "必填", "原始"),
            ("note", "备注", "文本", "成本项目与得率假设需保持一致", "选填", "原始"),
        ],
        "formulas": {
            9: lambda row: f'=IF(OR(D{row}="",E{row}="",F{row}="",G{row}=""),"",D{row}*F{row}+E{row}*G{row})',
            10: lambda row: f'=IF(OR(I{row}="",C{row}="",H{row}=""),"",I{row}-C{row}-H{row})',
        },
    },
]

SOURCE_ROWS = [
    ("USDA PSD", "全球油籽和豆粕产量、贸易、压榨、消费、库存", "年度/每月修订", "千吨、千公顷", "营销年度", "https://apps.fas.usda.gov/PSDOnlineDataServices/swagger/ui/index"),
    ("USDA NASS", "美国大豆面积、单产、产量", "年度/月度修订", "英亩、蒲式耳/英亩", "自然年/作物年度", "https://quickstats.nass.usda.gov/"),
    ("CONAB", "巴西大豆面积、单产、产量", "年度/月度修订", "千公顷、千吨", "巴西作物年度", "https://www.conab.gov.br/info-agro/safras"),
    ("阿根廷农业部", "阿根廷大豆面积、单产、产量", "年度/月度修订", "公顷、吨", "阿根廷作物年度", "https://www.magyp.gob.ar/sitio/areas/estimaciones/"),
    ("中国海关总署", "大豆进口数量、金额、来源国", "月度", "万吨、亿元或吨、美元", "自然月", "https://english.customs.gov.cn/Statistics/Statistics"),
    ("中国饲料工业协会", "分品种饲料产量", "月度/年度", "万吨", "自然月", "https://www.chinafeed.org.cn/"),
    ("国家统计局", "生猪存栏、出栏等", "季度/年度", "万头", "季度末/年末", "https://data.stats.gov.cn/"),
    ("农业农村部", "能繁母猪、屠宰量、养殖利润等", "月度", "万头、元/头", "自然月", "https://www.moa.gov.cn/ztzl/szcpxx/jdsj/"),
    ("大连商品交易所", "豆粕期货和交割信息", "日度汇总为月度", "元/吨、手", "月均/月末", "https://www.dce.com.cn/"),
    ("行业资讯供应商", "到港、压榨、油厂库存、基差", "周度汇总为月度", "万吨、元/吨", "月合计/月均/月末", "请填写供应商或终端名称"),
]


def set_title(ws, last_col: int, text: str) -> None:
    end = get_column_letter(last_col)
    ws.merge_cells(f"A1:{end}2")
    cell = ws["A1"]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=RED)
    cell.font = Font(name=FONT_NAME, size=17, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 8
    ws.sheet_view.showGridLines = False


def style_header(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 34


def style_input_grid(ws, first_row: int, last_row: int, last_col: int) -> None:
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
            cell.font = Font(name=FONT_NAME, size=9, color=BLUE)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    for row in range(first_row, last_row + 1):
        ws.row_dimensions[row].height = 22


def create_guide(ws) -> None:
    set_title(ws, 8, "豆粕研究数据采集模板｜月度与年度")
    ws.append([])
    ws["A4"], ws["B4"], ws["C4"] = "规则", "要求", "原因"
    ws.merge_cells("C4:H4")
    style_header(ws, 4, 1, 8)
    rules = [
        ("日期", "月度数据统一填当月1日，例如2026-01-01，显示格式为yyyy-mm；年度数据使用营销年度YYYY/YY。", "便于排序、合并和作图"),
        ("数值", "只填数字，不要在单元格里附加“万吨”“%”等文字；单位已写在字段名和字段字典中。", "保证可计算"),
        ("百分比", "Excel中填15.2%或0.152，不要填15.2。", "避免放大100倍"),
        ("缺失值", "没有数据时保持空白，不填0、不填“-”。", "0代表真实的零"),
        ("来源", "每行保留来源网址；付费终端数据可填终端名称和发布日期。", "保证可追溯"),
        ("版本", "USDA等预测会修订，同一营销年度不同版本不得覆盖；用version_date区分。", "保留预测变化"),
        ("月度汇总", "价格填月均及月末；库存填月末；进口、压榨和产量填月度合计。", "统一统计口径"),
        ("颜色", "浅黄色和蓝字为需要采集填写；浅绿色为公式自动计算。", "降低误填风险"),
    ]
    for row_no, values in enumerate(rules, start=5):
        for col_no, value in enumerate(values, start=1):
            ws.cell(row_no, col_no, value)
            ws.cell(row_no, col_no).font = Font(name=FONT_NAME, size=10, color=DARK)
            ws.cell(row_no, col_no).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_no].height = 34

    headers = ["工作表", "频率", "采集内容", "首选来源", "必须先采", "输出用途"]
    for col, value in enumerate(headers, start=1):
        ws.cell(14, col, value)
    ws.merge_cells("F14:H14")
    style_header(ws, 14, 1, 8)
    outputs = [
        "全球油料分布", "主产国面积、单产和产量", "中国进口需求", "国内豆粕供给与库存",
        "豆粕需求预测", "豆粕需求预测", "基差、季节性与价格强弱", "成本与利润",
    ]
    must_collect = {"01全球油籽年度", "02大豆主产国年度", "03中国进口月度", "04压榨库存月度", "05饲料月度", "06养殖月度", "07价格基差月度"}
    for index, schema in enumerate(SCHEMAS, start=15):
        values = [
            schema["name"], schema["frequency"], schema["title"].split("｜")[0],
            schema["source"], "是" if schema["name"] in must_collect else "否",
            outputs[index - 15],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(index, col, value)
            ws.cell(index, col).font = Font(name=FONT_NAME, size=9, color=DARK)
            ws.cell(index, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.conditional_formatting.add(
        f"E15:E{14 + len(SCHEMAS)}",
        FormulaRule(formula=["E15=\"是\""], fill=PatternFill("solid", fgColor=PALE_RED), font=Font(bold=True, color="9C0006")),
    )
    widths = {"A": 20, "B": 28, "C": 30, "D": 31, "E": 13, "F": 25, "G": 18, "H": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_dictionary(ws) -> None:
    set_title(ws, 9, "字段字典｜数据格式、单位与必填要求")
    headers = ["工作表", "字段名", "中文名称", "数据类型", "格式/单位", "是否必填", "频率", "原始/公式", "说明"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, 9)
    row = 5
    for schema in SCHEMAS:
        for field in schema["fields"]:
            values = [schema["name"], *field[:5], schema["frequency"], field[5], schema["title"]]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                cell.font = Font(name=FONT_NAME, size=9, color=DARK)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
    ws.freeze_panes = "D5"
    widths = [22, 31, 23, 13, 34, 13, 19, 13, 42]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_sources(ws) -> None:
    set_title(ws, 6, "数据源与统一口径")
    headers = ["数据源", "需要采集的内容", "建议频率", "统一单位", "时间口径", "网址"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, 6)
    for row_no, values in enumerate(SOURCE_ROWS, start=5):
        for col_no, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col_no, value)
            cell.font = Font(name=FONT_NAME, size=9, color=DARK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    widths = [24, 43, 20, 24, 25, 52]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_input_sheet(ws, schema: dict, table_index: int) -> None:
    field_count = len(schema["fields"])
    set_title(ws, field_count, schema["title"])
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=field_count)
    source_cell = ws.cell(3, 1)
    source_cell.value = f'首选来源：{schema["source"]}　{schema["source_url"]}'
    source_cell.fill = PatternFill("solid", fgColor=PALE_RED)
    source_cell.font = Font(name=FONT_NAME, size=9, italic=True, color="9C0006")

    for col, field in enumerate(schema["fields"], start=1):
        ws.cell(4, col, f"{field[1]}\n{field[4]}\n{field[5]}")
    style_header(ws, 4, 1, field_count)
    style_input_grid(ws, 5, MAX_ROWS, field_count)

    formula_columns = schema.get("formulas", {})
    for col, field in enumerate(schema["fields"], start=1):
        letter = get_column_letter(col)
        width = 24 if field[2] == "文本" else 14 if field[2] == "日期" else 17
        ws.column_dimensions[letter].width = width
        for row in range(5, MAX_ROWS + 1):
            cell = ws.cell(row, col)
            if field[2] == "日期":
                cell.number_format = "yyyy-mm"
            elif field[2] == "百分比":
                cell.number_format = "0.0%"
            elif field[2] == "数值":
                cell.number_format = "#,##0.0"
            if field[5] == "公式":
                cell.fill = PatternFill("solid", fgColor=FORMULA_FILL)
                cell.font = Font(name=FONT_NAME, size=9, color=DARK)
        if col in formula_columns:
            for row in range(5, MAX_ROWS + 1):
                ws.cell(row, col, formula_columns[col](row))

    table_ref = f"A4:{get_column_letter(field_count)}{MAX_ROWS}"
    table = Table(displayName=f"DataTable{table_index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "C5"


def validate_workbook(path: Path) -> None:
    check = load_workbook(path, data_only=False)
    expected = ["00使用说明", "字段字典", "数据源与口径", *[s["name"] for s in SCHEMAS]]
    if check.sheetnames != expected:
        raise ValueError(f"工作表不一致：{check.sheetnames}")
    for ws in check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "#REF!" in cell.value:
                    raise ValueError(f"发现无效引用：{ws.title}!{cell.coordinate}")
    if not str(check["08成本利润月度"]["I5"].value).startswith("="):
        raise ValueError("成本利润公式缺失")


def main() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("00使用说明")
    dictionary = wb.create_sheet("字段字典")
    source_sheet = wb.create_sheet("数据源与口径")
    create_guide(guide)
    create_dictionary(dictionary)
    create_sources(source_sheet)

    for index, schema in enumerate(SCHEMAS, start=1):
        create_input_sheet(wb.create_sheet(schema["name"]), schema, index)

    wb.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)
    print(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    main()
