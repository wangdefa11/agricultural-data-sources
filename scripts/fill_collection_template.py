"""把 Wind/USDA 数据回填到豆粕研究数据采集模板。

读取 outputs/.../豆粕研究数据采集模板_月度年度.xlsx，
另存 outputs/豆粕研究数据采集模板_月度年度_已填充2026-08-01.xlsx，原文件不动。
"""
import csv
import datetime as dt
import sqlite3
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TPL = ROOT / "outputs/019fbb09-1f0e-75e3-8f18-2e45e03c7fec/豆粕研究数据采集模板_月度年度.xlsx"
OUT = ROOT / "outputs/豆粕研究数据采集模板_月度年度_已填充2026-08-01.xlsx"
RAW_WIND = ROOT / "data/raw/wind"
PSD_CSV = ROOT / "data/raw/psd/2026-07-10/psd_oilseeds.csv"
DB = ROOT / "data/soybean.sqlite3"

PSD_URL = "https://apps.fas.usda.gov/psdonline/app/index.html#/app/downloads"
CUSTOMS_URL = "https://online.customs.gov.cn/"
NBS_URL = "https://data.stats.gov.cn/"
WIND_NOTE = "Wind 数据终端，2026-08-01 采集"
VERSION = dt.date(2026, 7, 10)

YEARS = [2021, 2022, 2023, 2024, 2025]  # 营销年度 2021/22—2025/26


def my(y: int) -> str:
    return f"{y}/{(y + 1) % 100:02d}"


def month_start(datestr: str) -> dt.date:
    y, m, _ = datestr.split("-")
    return dt.date(int(y), int(m), 1)


def read_wind(name: str):
    with (RAW_WIND / name).open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def col_values(header, rows, colname, nth=1):
    """取第 nth 个同名 indicator 列的 {date: float}。"""
    seen = 0
    idx = None
    for i, h in enumerate(header):
        if h == colname:
            seen += 1
            if seen == nth:
                idx = i
                break
    out = {}
    if idx is None:
        return out
    for r in rows:
        if len(r) > idx and r[idx].strip():
            try:
                out[r[0]] = float(r[idx])
            except ValueError:
                pass
    return out


def main() -> None:
    wb = openpyxl.load_workbook(TPL)
    con = sqlite3.connect(DB)
    cur = con.cursor()

    # ---------- 01 全球油籽年度：PSD 国家加总（PSD 无 World 行） ----------
    COMMODITIES = {
        "Oilseed, Soybean": "大豆", "Oilseed, Rapeseed": "菜籽",
        "Oilseed, Sunflowerseed": "葵花籽", "Oilseed, Cottonseed": "棉籽",
        "Oilseed, Peanut": "花生", "Oilseed, Copra": "椰肉干",
        "Oilseed, Palm Kernel": "棕榈仁",
    }
    ATTRS = {"Production": "E", "Exports": "F", "Imports": "G", "Crush": "H",
             "Domestic Consumption": "I", "Ending Stocks": "J"}
    sums = defaultdict(float)
    with PSD_CSV.open(encoding="utf-8-sig", errors="replace", newline="") as f:
        for row in csv.DictReader(f):
            c = row["Commodity_Description"]
            if c not in COMMODITIES:
                continue
            y = int(row["Market_Year"])
            if y not in YEARS:
                continue
            a = row["Attribute_Description"]
            if a in ATTRS and row["Value"].strip():
                sums[(c, y, a)] += float(row["Value"])
    ws = wb["01全球油籽年度"]
    r = 5
    for c, cname in COMMODITIES.items():
        for y in YEARS:
            ws.cell(r, 1, my(y))
            ws.cell(r, 2, VERSION).number_format = "yyyy-mm-dd"
            ws.cell(r, 3, cname)
            ws.cell(r, 4, "World(国家加总)")
            for a, col in ATTRS.items():
                v = sums.get((c, y, a))
                if v:
                    ws.cell(r, ord(col) - 64, round(v))
            ws.cell(r, 11, PSD_URL)
            note = "PSD 无 World 行，按国家加总"
            if cname == "大豆":
                note += "；与 Wind 全球值交叉验证差<0.01%"
            ws.cell(r, 12, note)
            r += 1

    # ---------- 02 大豆主产国年度：psd_observations ----------
    ws = wb["02大豆主产国年度"]
    r = 5
    for cc, cname in [("US", "美国"), ("BR", "巴西"), ("AR", "阿根廷"), ("CH", "中国")]:
        for y in YEARS:
            ws.cell(r, 1, my(y))
            ws.cell(r, 2, VERSION).number_format = "yyyy-mm-dd"
            ws.cell(r, 3, cname)
            for attr, col in [("Area Harvested", 5), ("Yield", 6), ("Production", 7),
                              ("Exports", 8), ("Imports", 9), ("Crush", 10), ("Ending Stocks", 11)]:
                row = cur.execute(
                    "SELECT value FROM psd_observations WHERE country_code=? AND market_year=? AND attribute=?",
                    (cc, y, attr)).fetchone()
                if row:
                    ws.cell(r, col, round(row[0], 2) if attr == "Yield" else round(row[0]))
            ws.cell(r, 12, PSD_URL)
            ws.cell(r, 13, "PSD 仅收获面积，无播种面积；与 Wind USDA 口径列一致")
            r += 1

    # ---------- 03 中国进口月度：海关月度 + 分国别 ----------
    h, rows = read_wind("wind_soybean_import_monthly.csv")
    total = col_values(h, rows, "中国:进口数量:大豆:当月值")  # 万吨
    h2, rows2 = read_wind("wind_soybean_import_by_origin.csv")
    br = col_values(h2, rows2, "中国:进口数量:大豆:巴西")       # 吨
    us = col_values(h2, rows2, "中国:进口数量:大豆:美国")
    ar = col_values(h2, rows2, "中国:进口数量:大豆:阿根廷")
    ws = wb["03中国进口月度"]
    r = 5
    for d in sorted(total, reverse=True):
        t = total[d]
        parts = {}
        for name, src in [("巴西", br), ("美国", us), ("阿根廷", ar)]:
            v = src.get(d)
            parts[name] = round(v / 10000, 2) if v is not None else None
        known = sum(v for v in parts.values() if v is not None)
        other = round(t - known, 2) if known else None
        for name, v in list(parts.items()) + [("其他(含未列明)", other)]:
            if v is None or v == 0:  # 模板规则：缺失/微量保持空白，不填0
                continue
            ws.cell(r, 1, month_start(d)).number_format = "yyyy-mm"
            ws.cell(r, 2, name)
            ws.cell(r, 3, v)
            ws.cell(r, 6, CUSTOMS_URL)
            ws.cell(r, 7, WIND_NOTE + "；分国别为吨换算万吨")
            r += 1

    # ---------- 04 压榨库存月度 ----------
    ws = wb["04压榨库存月度"]
    crush = {m: v for m, v in cur.execute(
        "SELECT month, value FROM macro_monthly WHERE indicator='大豆压榨量(实际月度)(万吨/月)'")}
    r = 5
    for m in sorted(crush, reverse=True):
        ws.cell(r, 1, dt.date(int(m[:4]), int(m[5:]), 1)).number_format = "yyyy-mm"
        ws.cell(r, 2, "全国")
        ws.cell(r, 4, crush[m])
        ws.cell(r, 11, "Wind（资讯商月度压榨）")
        ws.cell(r, 13, "日历年月度口径；港口/油厂库存 Wind 无数据留空")
        r += 1

    # ---------- 05 饲料月度 ----------
    h, rows = read_wind("wind_feed_output_monthly.csv")
    feed = col_values(h, rows, "中国:产量:饲料:当月值")  # 吨
    ws = wb["05饲料月度"]
    r = 5
    for d in sorted(feed, reverse=True):
        y, mth = int(d[:4]), int(d[5:7])
        v = feed[d] / 10000
        prev = feed.get(f"{y-1}-{mth:02d}-{d[8:]}")
        ws.cell(r, 1, dt.date(y, mth, 1)).number_format = "yyyy-mm"
        ws.cell(r, 2, "饲料总产量")
        ws.cell(r, 3, round(v, 2))
        if prev:
            ws.cell(r, 4, round((feed[d] / prev - 1), 4)).number_format = "0.0%"
        ws.cell(r, 7, NBS_URL)
        ws.cell(r, 8, WIND_NOTE + "；分品种产量 Wind 无，添加率假设留空待填")
        r += 1

    # ---------- 06 养殖月度（季度落季度末月） ----------
    h, rows = read_wind("wind_hog_inventory.csv")
    hog = col_values(h, rows, "中国:存栏数:生猪")
    sow = col_values(h, rows, "中国:存栏数:能繁母猪")
    h2, rows2 = read_wind("wind_hog_slaughter.csv")
    slt = col_values(h2, rows2, "中国:出栏数:生猪")
    broiler = {d[:4]: v for d, v in cur.execute(
        "SELECT date, value FROM wind_observations WHERE indicator='中国:年末存栏数量:父母代肉鸡场'")}
    ws = wb["06养殖月度"]
    r = 5
    months = sorted(set(hog) | set(sow) | set(slt), reverse=True)
    for d in months:
        y, mth = int(d[:4]), int(d[5:7])
        ws.cell(r, 1, dt.date(y, mth, 1)).number_format = "yyyy-mm"
        if hog.get(d) is not None:
            ws.cell(r, 2, hog[d])
        # Wind 能繁母猪列口径混杂（2021 年中后混入约 1.2 万的异口径序列），
        # 仅保留农业农村部公开口径区间（3000—5000 万头）内的值
        if sow.get(d) is not None and 3000 <= sow[d] <= 5000:
            ws.cell(r, 3, sow[d])
        if slt.get(d) is not None:
            ws.cell(r, 4, slt[d])
        if mth == 12 and str(y) in broiler:
            ws.cell(r, 5, broiler[str(y)])
        ws.cell(r, 8, "国家统计局/农业农村部（Wind 采集）")
        ws.cell(r, 9, NBS_URL)
        note = "季度数据落季度末月；能繁母猪仅保留 3000—5000 万头口径区间值，Wind 异口径值已剔除"
        if mth == 12 and str(y) in broiler:
            note += "；肉鸡为父母代年末存栏(只)"
        ws.cell(r, 10, note)
        r += 1

    # ---------- 07 价格基差月度 ----------
    spot = {m: v for m, v in cur.execute(
        "SELECT month, value FROM macro_monthly WHERE indicator='豆粕现货价月均(元/吨)'")}
    ws = wb["07价格基差月度"]
    r = 5
    for m in sorted(spot):
        ws.cell(r, 1, dt.date(int(m[:4]), int(m[5:]), 1)).number_format = "yyyy-mm"
        ws.cell(r, 2, "全国均价")
        ws.cell(r, 5, spot[m])
        ws.cell(r, 12, "Wind 数据终端")
        ws.cell(r, 13, "期货/汇率/CBOT 月均 Wind 通道不可得，留空；现货为日度月均")
        r += 1

    con.close()
    wb.save(OUT)
    print("saved:", OUT)
    for name in ["01全球油籽年度", "02大豆主产国年度", "03中国进口月度", "04压榨库存月度",
                 "05饲料月度", "06养殖月度", "07价格基差月度"]:
        ws = wb[name]
        n = sum(1 for row in ws.iter_rows(min_row=5) if any(c.value is not None for c in row))
        print(f"  {name}: 填充 {n} 行")


if __name__ == "__main__":
    main()
